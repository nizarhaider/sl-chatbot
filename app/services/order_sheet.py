from __future__ import annotations

import json
import logging
import os
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class OrderLine:
    name: str
    quantity: int = 1
    sku: str = ""
    price: str = ""


@dataclass(frozen=True)
class CustomerOrder:
    customer_phone: str
    customer_message: str
    lines: list[OrderLine]

    def summary(self) -> str:
        return "\n".join(
            f"- {line.quantity} x {line.name}"
            f"{f' ({line.sku})' if line.sku else ''}"
            f"{f' - {line.price}' if line.price else ''}"
            for line in self.lines
        )


class OrderSheetClient:
    def __init__(self):
        self.spreadsheet_id = os.environ.get("GOOGLE_ORDERS_SPREADSHEET_ID")
        self.local_path = os.environ.get("LOCAL_ORDERS_PATH", "data/orders.xlsx")
        self.worksheet_name = os.environ.get("GOOGLE_ORDERS_WORKSHEET_NAME", "Orders")
        self.timezone_name = os.environ.get("ORDER_TIMEZONE", "Asia/Colombo")

    def append_order(self, order: CustomerOrder) -> bool:
        if not self.spreadsheet_id:
            return self._append_order_locally(order)

        try:
            service = self._sheets_service()
            values = self._rows_for_order(order)
            service.spreadsheets().values().append(
                spreadsheetId=self.spreadsheet_id,
                range=f"{self.worksheet_name}!A:G",
                valueInputOption="USER_ENTERED",
                insertDataOption="INSERT_ROWS",
                body={"values": values},
            ).execute()
            return True
        except Exception as exc:
            logger.error("Failed to append order to Google Sheet: %s", exc)
            return False

    def _append_order_locally(self, order: CustomerOrder) -> bool:
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except ImportError:
            logger.error("openpyxl is required to write local order workbooks")
            return False

        try:
            path = Path(self.local_path)
            path.parent.mkdir(parents=True, exist_ok=True)
            headers = [
                "created_at",
                "customer_phone",
                "product",
                "sku",
                "quantity",
                "price",
                "customer_message",
            ]

            if path.exists():
                workbook = load_workbook(path)
                worksheet = workbook[self.worksheet_name] if self.worksheet_name in workbook.sheetnames else workbook.active
                worksheet.title = self.worksheet_name
                if worksheet.max_row == 1 and not worksheet["A1"].value:
                    worksheet.append(headers)
            else:
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = self.worksheet_name
                worksheet.append(headers)

            for row in self._rows_for_order(order):
                worksheet.append(row)

            worksheet.freeze_panes = "A2"
            widths = {
                "A": 24,
                "B": 18,
                "C": 34,
                "D": 16,
                "E": 12,
                "F": 14,
                "G": 55,
            }
            for column, width in widths.items():
                worksheet.column_dimensions[column].width = width

            table_ref = f"A1:G{worksheet.max_row}"
            if "OrdersTable" not in worksheet.tables and worksheet.max_row >= 2:
                table = Table(displayName="OrdersTable", ref=table_ref)
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                worksheet.add_table(table)
            elif "OrdersTable" in worksheet.tables:
                worksheet.tables["OrdersTable"].ref = table_ref

            workbook.save(path)
            logger.info(
                "Appended %s order line(s) to local workbook %s; rows=%s",
                len(order.lines),
                path,
                worksheet.max_row,
            )
            return True
        except Exception as exc:
            logger.error("Failed to append order to local workbook: %s", exc)
            return False

    def _rows_for_order(self, order: CustomerOrder) -> list[list[str | int]]:
        created_at = datetime.now(ZoneInfo(self.timezone_name)).isoformat(timespec="seconds")
        return [
            [
                created_at,
                order.customer_phone,
                line.name,
                line.sku,
                line.quantity,
                line.price,
                order.customer_message,
            ]
            for line in order.lines
        ]

    def _sheets_service(self):
        try:
            from google.oauth2 import service_account
            from googleapiclient.discovery import build
        except ImportError as exc:
            raise RuntimeError("google-api-python-client and google-auth are required") from exc

        credentials = _service_account_credentials(
            scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
        return build("sheets", "v4", credentials=credentials, cache_discovery=False)


def _service_account_credentials(scopes: list[str]):
    from google.oauth2 import service_account

    json_path = os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE")
    json_blob = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")

    if json_path:
        return service_account.Credentials.from_service_account_file(json_path, scopes=scopes)
    if json_blob:
        return service_account.Credentials.from_service_account_info(json.loads(json_blob), scopes=scopes)
    raise RuntimeError("Set GOOGLE_SERVICE_ACCOUNT_FILE or GOOGLE_SERVICE_ACCOUNT_JSON")
