from __future__ import annotations

import csv
import logging
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class Product:
    name: str
    sku: str = ""
    price: str = ""
    stock: str = ""
    description: str = ""
    variant: str = ""
    image_url: str = ""
    product_url: str = ""

    def display_name(self) -> str:
        if self.variant and self.variant.lower() not in self.name.lower():
            return f"{self.name} - {self.variant}"
        return self.name


class ProductCatalog:
    def __init__(self, path: str | None = None):
        self.path = path or os.environ.get("PRODUCT_CATALOG_PATH", "data/products.xlsx")
        self._products: list[Product] = []
        self._loaded_mtime: float | None = None

    def search(self, query: str, limit: int = 5) -> list[Product]:
        products = self.load()
        tokens = _tokens(query)
        if not tokens:
            return []

        scored: list[tuple[int, Product]] = []
        for product in products:
            haystack = " ".join(
                [
                    product.name,
                    product.sku,
                    product.price,
                    product.stock,
                    product.description,
                    product.variant,
                    product.image_url,
                    product.product_url,
                ]
            ).lower()
            score = sum(3 if token in product.name.lower() else 1 for token in tokens if token in haystack)
            if score:
                scored.append((score, product))

        scored.sort(key=lambda item: item[0], reverse=True)
        return [product for _, product in scored[:limit]]

    def get_by_sku(self, sku: str) -> Product | None:
        normalized_sku = sku.strip().lower()
        if not normalized_sku:
            return None

        for product in self.load():
            if product.sku.strip().lower() == normalized_sku:
                return product
        return None

    def load(self) -> list[Product]:
        path = Path(self.path)
        if not path.exists():
            logger.warning("Product catalog file does not exist: %s", path)
            self._products = []
            self._loaded_mtime = None
            return []

        mtime = path.stat().st_mtime
        if self._loaded_mtime == mtime:
            return self._products

        if path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            rows = self._read_xlsx(path)
        elif path.suffix.lower() == ".csv":
            rows = self._read_csv(path)
        else:
            logger.error("Unsupported product catalog format: %s", path.suffix)
            rows = []

        self._products = [_product_from_row(row) for row in rows if _row_value(row, "name", "product", "item")]
        self._loaded_mtime = mtime
        logger.info("Loaded %s products from %s", len(self._products), path)
        return self._products

    def _read_csv(self, path: Path) -> list[dict[str, str]]:
        with path.open(newline="", encoding="utf-8-sig") as file:
            reader = csv.DictReader(file)
            return [_normalize_row(row) for row in reader]

    def _read_xlsx(self, path: Path) -> list[dict[str, str]]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl is required to read Excel product catalogs")
            return []

        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet_name = os.environ.get("PRODUCT_CATALOG_SHEET")
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return []

        normalized_headers = [_normalize_key(str(header or "")) for header in headers]
        records: list[dict[str, str]] = []
        for row in rows:
            record = {}
            for index, value in enumerate(row):
                if index < len(normalized_headers) and normalized_headers[index]:
                    record[normalized_headers[index]] = "" if value is None else str(value).strip()
            records.append(record)
        return records


def _product_from_row(row: dict[str, str]) -> Product:
    return Product(
        name=_row_value(row, "name", "product", "item", "product_name"),
        sku=_row_value(row, "sku", "code", "item_code"),
        price=_row_value(row, "price", "selling_price", "unit_price"),
        stock=_row_value(row, "stock", "qty", "quantity", "inventory", "available"),
        description=_row_value(row, "description", "details", "notes"),
        variant=_row_value(row, "variant", "shade", "size", "colour", "color"),
        image_url=_row_value(row, "image_url", "image", "photo_url", "photo"),
        product_url=_row_value(row, "product_url", "url", "link"),
    )


def _normalize_row(row: dict[str, str]) -> dict[str, str]:
    return {_normalize_key(key): str(value or "").strip() for key, value in row.items()}


def _normalize_key(key: str) -> str:
    return key.strip().lower().replace(" ", "_").replace("-", "_")


def _row_value(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value:
            return value
    return ""


def _tokens(text: str) -> list[str]:
    stop_words = {
        "a",
        "an",
        "and",
        "any",
        "available",
        "buy",
        "can",
        "do",
        "for",
        "have",
        "i",
        "in",
        "is",
        "it",
        "me",
        "need",
        "of",
        "order",
        "price",
        "product",
        "stock",
        "the",
        "want",
        "what",
        "with",
        "you",
    }
    return [
        token
        for token in "".join(char.lower() if char.isalnum() else " " for char in text).split()
        if len(token) > 2 and token not in stop_words
    ]
