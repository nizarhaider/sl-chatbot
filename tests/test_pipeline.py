import asyncio
import os
import logging
import tempfile
import unittest
from unittest.mock import patch

import httpx

os.environ.setdefault("GOOGLE_API_KEY", "test-key")

from app.main import app
from app.chat_agent.agent import ChatAgent, ChatAgentResult
from app.services.order_sheet import CustomerOrder, OrderLine
from app.services.product_catalog import Product
from app.voice_agent.agent import RealtimeAudioTrack
from app.voice_agent.gemini_turn_pipeline import GeminiTurnPipeline


class WebhookPipelineTests(unittest.IsolatedAsyncioTestCase):
    async def asyncSetUp(self):
        self.transport = httpx.ASGITransport(app=app)
        self.client = httpx.AsyncClient(transport=self.transport, base_url="http://testserver")

    async def asyncTearDown(self):
        await self.client.aclose()

    async def test_call_connect_webhook_schedules_offer_handling(self):
        handled = asyncio.Event()
        captured = {}

        async def fake_handle_offer(call_id: str, sdp_offer: str, caller_phone: str = ""):
            captured["call_id"] = call_id
            captured["sdp_offer"] = sdp_offer
            captured["caller_phone"] = caller_phone
            handled.set()

        payload = {
            "object": "whatsapp_business_account",
            "entry": [
                {
                    "changes": [
                        {
                            "value": {
                                "calls": [
                                    {
                                        "event": "connect",
                                        "id": "call-123",
                                        "from": "94770000000",
                                        "session": {
                                            "sdp_type": "offer",
                                            "sdp": "v=0\r\ns=test\r\n",
                                        },
                                    }
                                ]
                            }
                        }
                    ]
                }
            ],
        }

        with patch("app.webhooks.whatsapp.webrtc_service.handle_offer", new=fake_handle_offer):
            response = await self.client.post("/webhook", json=payload)
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.text, "EVENT_RECEIVED")
            await asyncio.wait_for(handled.wait(), timeout=1)

        self.assertEqual(
            captured,
            {
                "call_id": "call-123",
                "sdp_offer": "v=0\r\ns=test\r\n",
                "caller_phone": "94770000000",
            },
        )

    async def test_text_message_webhook_sends_chat_agent_reply(self):
        sent = asyncio.Event()
        captured = {}

        async def fake_process_message(text: str, sender_id: str | None = None) -> ChatAgentResult:
            captured["prompt"] = text
            captured["sender_id"] = sender_id
            return ChatAgentResult("Hello from the bot")

        async def fake_send_message(to: str, text: str) -> bool:
            captured["to"] = to
            captured["reply"] = text
            sent.set()
            return True

        payload = {
            "object": "whatsapp_business_account",
            "entry": [
                {
                    "changes": [
                        {
                            "value": {
                                "messages": [
                                    {
                                        "type": "text",
                                        "from": "94770000000",
                                        "text": {"body": "Hi"},
                                    }
                                ]
                            }
                        }
                    ]
                }
            ],
        }

        with patch("app.webhooks.whatsapp.chat_agent.process_message", new=fake_process_message):
            with patch("app.webhooks.whatsapp.whatsapp_api.send_message", new=fake_send_message):
                response = await self.client.post("/webhook", json=payload)
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.text, "EVENT_RECEIVED")
                await asyncio.wait_for(sent.wait(), timeout=1)

        self.assertEqual(
            captured,
            {
                "prompt": "Hi",
                "sender_id": "94770000000",
                "to": "94770000000",
                "reply": "Hello from the bot",
            },
        )

    async def test_text_message_webhook_sends_product_images_from_agent_result(self):
        sent = asyncio.Event()
        captured = {"messages": [], "images": []}

        async def fake_process_message(text: str, sender_id: str | None = None) -> ChatAgentResult:
            return ChatAgentResult(
                "Here is the product.",
                image_urls=["https://img.drz.lazcdn.com/static/lk/p/example.jpg"],
            )

        async def fake_send_message(to: str, text: str) -> bool:
            captured["messages"].append((to, text))
            return True

        async def fake_send_image(to: str, image_url: str, caption: str = "") -> bool:
            captured["images"].append((to, image_url, caption))
            sent.set()
            return True

        payload = {
            "object": "whatsapp_business_account",
            "entry": [
                {
                    "changes": [
                        {
                            "value": {
                                "messages": [
                                    {
                                        "type": "text",
                                        "from": "94770000000",
                                        "text": {"body": "Show me nails"},
                                    }
                                ]
                            }
                        }
                    ]
                }
            ],
        }

        with patch("app.webhooks.whatsapp.chat_agent.process_message", new=fake_process_message):
            with patch("app.webhooks.whatsapp.whatsapp_api.send_message", new=fake_send_message):
                with patch("app.webhooks.whatsapp.whatsapp_api.send_image", new=fake_send_image):
                    response = await self.client.post("/webhook", json=payload)
                    self.assertEqual(response.status_code, 200)
                    await asyncio.wait_for(sent.wait(), timeout=1)

        self.assertEqual(captured["messages"], [("94770000000", "Here is the product.")])
        self.assertEqual(
            captured["images"],
            [("94770000000", "https://img.drz.lazcdn.com/static/lk/p/example.jpg", "")],
        )

    async def test_confirmed_order_updates_sheet_and_notifies_manager(self):
        sent = asyncio.Event()
        replies = []
        manager_messages = []
        captured_orders = []

        async def fake_send_message(to: str, text: str) -> bool:
            if to == "94742530708":
                manager_messages.append(text)
            else:
                replies.append((to, text))
            if manager_messages:
                sent.set()
            return True

        def fake_append_order(order: CustomerOrder) -> bool:
            captured_orders.append(order)
            return True

        app_chat_agent = __import__("app.webhooks.whatsapp", fromlist=["chat_agent"]).chat_agent
        app_chat_agent.pending_orders["94770000000"] = CustomerOrder(
            customer_phone="94770000000",
            customer_message="Order Glow Serum qty 2",
            lines=[OrderLine(name="Glow Serum", quantity=2, sku="GS-01", price="2500")],
        )

        payload = {
            "object": "whatsapp_business_account",
            "entry": [
                {
                    "changes": [
                        {
                            "value": {
                                "messages": [
                                    {
                                        "type": "text",
                                        "from": "94770000000",
                                        "text": {"body": "confirm"},
                                    }
                                ]
                            }
                        }
                    ]
                }
            ],
        }

        with patch.object(app_chat_agent.order_sheet, "append_order", new=fake_append_order):
            with patch("app.webhooks.whatsapp.whatsapp_api.send_message", new=fake_send_message):
                response = await self.client.post("/webhook", json=payload)
                self.assertEqual(response.status_code, 200)
                await asyncio.wait_for(sent.wait(), timeout=1)

        self.assertEqual(captured_orders[0].customer_phone, "94770000000")
        self.assertIn("confirmed", replies[0][1].lower())
        self.assertIn("New confirmed WhatsApp order", manager_messages[0])
        self.assertIn("Glow Serum", manager_messages[0])

    async def test_product_query_uses_catalog_matches(self):
        agent = ChatAgent()

        with patch.object(
            agent.product_catalog,
            "search",
            return_value=[
                Product(
                    name="Glow Serum",
                    sku="GS-01",
                    price="2500",
                    stock="12",
                    image_url="https://img.drz.lazcdn.com/static/lk/p/glow.jpg",
                )
            ],
        ):
            result = await agent.process_message("Do you have glow serum price?", "94770000000")

        self.assertIn("Glow Serum", result.reply)
        self.assertIn("2500", result.reply)
        self.assertEqual(result.image_urls, ["https://img.drz.lazcdn.com/static/lk/p/glow.jpg"])

    async def test_exact_order_request_creates_single_pending_line(self):
        agent = ChatAgent()

        with patch.object(
            agent.product_catalog,
            "search",
            return_value=[
                Product(name="Classic Nude Press-On Set", sku="NBA-NUDE", price="Confirm with store"),
                Product(name="French Tip Press-On Set", sku="NBA-FRENCH", price="Confirm with store"),
                Product(name="Nails By Ayidaah Press-On Nail Set", sku="NBA-PRESSON-24", price="1690"),
            ],
        ):
            result = await agent.process_message(
                "I want the classic nude press on set",
                "94770000000",
            )

        self.assertIn("Classic Nude Press-On Set", result.reply)
        self.assertIn("CONFIRM", result.reply)
        self.assertEqual(len(agent.pending_orders["94770000000"].lines), 1)
        self.assertEqual(agent.pending_orders["94770000000"].lines[0].sku, "NBA-NUDE")

    async def test_generic_order_request_asks_customer_to_choose_one_item(self):
        agent = ChatAgent()

        with patch.object(
            agent.product_catalog,
            "search",
            return_value=[
                Product(name="Classic Nude Press-On Set", sku="NBA-NUDE", price="Confirm with store"),
                Product(name="French Tip Press-On Set", sku="NBA-FRENCH", price="Confirm with store"),
            ],
        ):
            result = await agent.process_message("I need to get nails", "94770000000")

        self.assertIn("Please reply with the number or SKU", result.reply)
        self.assertNotIn("94770000000", agent.pending_orders)

    async def test_customer_can_select_last_search_result_by_number(self):
        agent = ChatAgent()
        agent.last_product_searches["94770000000"] = [
            Product(name="Classic Nude Press-On Set", sku="NBA-NUDE", price="Confirm with store"),
            Product(name="French Tip Press-On Set", sku="NBA-FRENCH", price="Confirm with store"),
        ]

        result = await agent.process_message("2", "94770000000")

        self.assertIn("French Tip Press-On Set", result.reply)
        self.assertEqual(agent.pending_orders["94770000000"].lines[0].sku, "NBA-FRENCH")

    async def test_order_sheet_appends_to_local_workbook_without_google_config(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            agent = ChatAgent()
            agent.order_sheet.spreadsheet_id = ""
            agent.order_sheet.local_path = os.path.join(temp_dir, "orders.xlsx")

            written = agent.order_sheet.append_order(
                CustomerOrder(
                    customer_phone="94770000000",
                    customer_message="Order French tips",
                    lines=[OrderLine(name="French Tip Press-On Set", quantity=1, sku="NBA-FRENCH")],
                )
            )

            self.assertTrue(written)

            from openpyxl import load_workbook

            workbook = load_workbook(agent.order_sheet.local_path)
            worksheet = workbook["Orders"]
            self.assertEqual(worksheet.max_row, 2)
            self.assertEqual(worksheet["C2"].value, "French Tip Press-On Set")

    async def test_default_chat_agent_prompt_is_ayidaah_service_agent(self):
        agent = ChatAgent()
        prompt = agent._build_system_prompt()

        self.assertIn("Ayidaah Beauty", prompt)
        self.assertIn("Monday to Saturday, 9:00 am to 8:00 pm", prompt)
        self.assertIn("closed on Sunday", prompt)
        self.assertIn("You do not have access to live inventory", prompt)
        self.assertIn("+94 77 167 9595", prompt)

    async def test_gemini_turn_pipeline_emits_greeting_through_tts(self):
        output_track = RealtimeAudioTrack()
        captured = {}
        pcm_bytes = b"\x01\x02" * 480

        class FakeRealtimeTTS:
            async def speak(self, text: str, on_audio_chunk):
                captured["text"] = text
                on_audio_chunk(pcm_bytes, 24000)
                return 0.02

            async def prewarm(self):
                captured["prewarmed"] = True

        class EndTrack:
            async def recv(self):
                raise RuntimeError("done")

        pipeline = GeminiTurnPipeline(
            prepare_tts_text=lambda text: text,
            interrupt_playback=lambda call_id, output_track: None,
            tts=FakeRealtimeTTS(),
        )

        with patch("app.voice_agent.gemini_turn_pipeline.TURN_GREETING_DELAY_SECONDS", 0):
            await pipeline.run(
                call_id="call-turn",
                input_track=EndTrack(),
                output_track=output_track,
                playback_generation={"call-turn": 0},
            )

        queued_audio = []
        while not output_track.queue.empty():
            queued_audio.append(output_track.queue.get_nowait())

        self.assertEqual(
            captured["text"],
            "Hello, this is Homelands. Please say Sinhala, Tamil, or English.",
        )
        self.assertTrue(queued_audio)

    async def test_gemini_turn_pipeline_speaks_prepared_text(self):
        output_track = RealtimeAudioTrack()
        captured = {}

        class FakeRealtimeTTS:
            async def speak(self, text: str, on_audio_chunk):
                captured["text"] = text
                return 0.02

            async def prewarm(self):
                return None

        pipeline = GeminiTurnPipeline(
            prepare_tts_text=lambda text: " ".join(text.split()),
            interrupt_playback=lambda call_id, output_track: None,
            tts=FakeRealtimeTTS(),
        )

        await pipeline._speak(
            "call-turn",
            "  Hello,   customer.  ",
            output_track,
            {"call-turn": 0},
        )

        self.assertEqual(captured["text"], "Hello, customer.")

    async def test_realtime_audio_track_resamples_mono_pcm_for_whatsapp(self):
        pcm_bytes = b"\x01\x02" * 480
        output_track = RealtimeAudioTrack()

        output_track.add_pcm_audio(pcm_bytes, 24000)

        queued_audio = []
        while not output_track.queue.empty():
            queued_audio.append(output_track.queue.get_nowait())

        self.assertTrue(queued_audio)
        self.assertEqual(output_track.sample_rate, 48000)
        self.assertGreaterEqual(len(b"".join(queued_audio)), len(pcm_bytes) * 2)


if __name__ == "__main__":
    unittest.main()
